"""
Service layer for returns app - handles CSV/Excel processing and import.
"""
import csv
import io
import json
import logging
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Any, Optional, Tuple

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.db.models import Q, Count, Sum

from .models import (
    ReturnTemplate, ReturnSubmission, ReturnData, 
    RegionalReturnSummary, StationReturnStatus,
    MonthlySubmissionTracker, ReturnTypeStatus
)
from prison.models import PrisonStation, Prisoner

logger = logging.getLogger(__name__)


class DefaultTemplateService:
    """
    Service for creating and managing default return templates.
    """

    # Template definitions for each category
    TEMPLATE_DEFINITIONS = {
        'convicted_inmates': {
            'name': 'Convicted Inmates Return',
            'description': 'Monthly return for convicted inmates held at a prison station.',
            'columns': [
                'SER.NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT/CASE.No', 'SENTENCE', 'DATE OF COMMITTAL',
                'EXPIRY DATE OF RELEASE WITHOUT REMISSION',
                'EXPIRY DATE OF RELEASE WITH REMISSION'
            ],
            'example_data': [
                ['1', 'ZA02/26', 'PATRICK PHIRI', 'V: Khata T/A: Malembo D. Kasungu', 'M', '26',
                 'Theft', 'F.G.M Mulungudzi Zomba 02/25', '72 months', '20-09-2025', '19-09-2031', '19-09-2029'],
            ]
        },
        'due_discharge': {
            'name': 'Due Discharge Return',
            'description': 'Monthly return for prisoners due for discharge.',
            'columns': [
                'S/No', 'PRIS.No', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT/CASE.No', 'SENTENCE', 'DATE OF CONVICTION',
                'AMNESTY EARNED', 'DUE DATE', 'REMARKS'
            ],
            'example_data': [
                ['1', 'ZA02/26', 'PATRICK PHIRI', 'V: Khata T/A: Malembo D. Kasungu', 'M', '26',
                 'Theft', 'F.G.M Mulungudzi Zomba 20/25', '72 months', '20-09-2025', '6 months', '04-04-2029', 'Loss of remission 15 days'],
            ]
        },
        'remand_murder': {
            'name': 'Remand Murder Prisoners Return',
            'description': 'Monthly return for remand murder prisoners.',
            'columns': [
                'Ser.No', 'PRIS.NO', 'NAME', 'PARTICULARS', 'AGE', 'SEX',
                'OFFENSE', 'COURT/CASE.No', 'DATE OF ADMISSION',
                'ARRESTING AUTHORITY', 'LAST COURT APPEARANCE',
                'JUDGE NAME', 'STATUS'
            ],
            'example_data': [
                ['1', 'ZA02/26', 'PATRICK PHIRI', 'V: Khata T/A: Malembo D. Kasungu', '26', 'M',
                 'Murder', 'HIGH COURT ZOMBA 02/25', '20-09-2025', 'Jali Police', '15-10-2025', 'Judge Mwambe', 'Pending'],
            ]
        },
        'convicted_foreigners': {
            'name': 'Convicted Foreigners Return',
            'description': 'Monthly return for convicted foreign nationals.',
            'columns': [
                'SER.NO', 'PRIS.No', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT', 'SENTENCE', 'DATE OF CONVICTION',
                'AMNESTY EARNED', 'EXPECTED DATE OF RELEASE'
            ],
            'example_data': [
                ['1', 'ZA04/26', 'John Zanda', 'V: P: D: C: Mozambique', 'M', '26',
                 'Theft', 'F.G.M C/No 2/25 Zomba', '72 months', '20-09-2025', 'Nil', '04-10-2029'],
                ['2', 'ZA22/26', 'Ethel Game', 'V: P: D: C: Zambia', 'F', '26',
                 'Theft', 'F.G.M C/No 230/25 Zomba', '72 months', '02-09-2025', 'Nil', '01-09-2029'],
            ]
        },
        'general_remand': {
            'name': 'General Remandees Return',
            'description': 'Monthly return for general remandees.',
            'columns': [
                'SER.NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT/CASE.No', 'DATE OF ADMISSION',
                'ARRESTING AUTHORITY', 'LAST COURT APPEARANCE', 'CASE STATUS'
            ],
            'example_data': [
                ['1', 'Mzu9/25', 'Tsedeke Girma', 'V: Hosana P: Hosana D: Hosana C: Zambia', 'M', '18',
                 'Illegal Entry', 'SGM C/No 14/25 Euthini', '22-02-2025', 'Kafukule Immigration', '20-03-2025', 'Waiting deportation'],
                ['2', 'Mzu10/25', 'James Phiri', 'V: Agoni P: Domwe D: Chitungwi C: Zimbabwe', 'M', '20',
                 'Theft', 'FGM C/No 102/25 Mzuzu', '22-02-2025', 'Mzuzu Police', '20-03-2025', 'Waiting for judgement'],
            ]
        },
        'pardon_consideration': {
            'name': 'Pardon Consideration List',
            'description': 'List of convicted prisoners to be considered for pardon.',
            'columns': [
                'NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT', 'SENTENCE', 'DOC', 'SENT WITH REM.',
                'SENT. SERVED', 'AMNESTY EARNED', 'EXPECTED DATE OF RELEASE',
                'PARTICULARS OF PREVIOUS CONVICTION', 'CONDUCT'
            ],
            'example_data': [
                ['1', 'BT 9/25', 'JUSTIN BANDA', 'V: CHIDE T/A: BVUMBWE D: THYOLO', 'M', '24',
                 'THEFT', 'F.G.M C/NO 785/25 MIDIMA', '9 MONTHS IHL', '02-07-2025', '6 MONTHS',
                 '3 MONTHS', '', '01-01-26', 'ONE BURLARY AND THEFT', 'GOOD'],
            ]
        },
        'chronically_ill': {
            'name': 'Chronically ill Convicted Inmates',
            'description': 'List of chronically ill convicted inmates proposed for pardon.',
            'columns': [
                'NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT', 'SENTENCE', 'DOC', 'SENT WITH REM.',
                'AMNESTY EARNED', 'EXPECTED DATE OF RELEASE',
                'PARTICULARS OF PREVIOUS CONVICTION', 'CONDUCT'
            ],
            'example_data': [
                ['1', 'BT 9/25', 'JUSTIN BANDA', 'V: CHIDE T/A: BVUMBWE D: THYOLO', 'M', '24',
                 'THEFT', 'F.G.M C/NO 785/25 MIDIMA', '30 MONTHS IHL', '02-07-2023', '20 MONTHS',
                 'TWO-THREE MONTHS', '01-06-2024', 'NIL', 'GOOD'],
            ]
        },
        'elderly_inmates': {
            'name': 'Elderly Inmates (70+)',
            'description': 'List of convicted elderly inmates aged 70 years and above proposed for pardon.',
            'columns': [
                'NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT', 'SENTENCE', 'DOC', 'SENT WITH REM.',
                'SENT. SERVED', 'AMNESTY EARNED', 'EXPECTED DATE OF RELEASE',
                'PARTICULARS OF PREVIOUS CONVICTION', 'CONDUCT'
            ],
            'example_data': [
                ['1', 'BT 9/25', 'JUSTIN BANDA', 'V: CHIDE T/A: BVUMBWE D: THYOLO', 'M', '72',
                 'THEFT', 'F.G.M C/NO 785/25 MIDIMA', '30 MONTHS IHL', '02-07-2023', '20 MONTHS',
                 'TWO-THREE MONTHS', 'SIX MONTHS', '01-06-2024', 'NIL', 'GOOD'],
            ]
        },
        'discharged_reduction': {
            'name': 'Discharged After Reduction',
            'description': 'Return for prisoners discharged after effecting 6 months reduction from their sentences.',
            'columns': [
                'NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT', 'SENTENCE', 'DOC', 'AMNESTY EARNED',
                'EXPECTED DATE OF RELEASE', 'EDR AFTER AMNESTY'
            ],
            'example_data': [
                ['1', 'BT 9/25', 'JUSTIN BANDA', 'V: CHIDE T/A: BVUMBWE D: THYOLO', 'M', '24',
                 'THEFT', 'F.G.M C/NO 785/25 MIDIMA', '30 MONTHS IHL', '02-07-2023', 'SIX MONTHS',
                 '01-06-2024', '31-12-2023'],
            ]
        },
        'children_with_mothers': {
            'name': 'Children Accompanying Mothers',
            'description': 'Return for children accompanying their mothers in prison.',
            'columns': [
                'S/N', 'PRI.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT CASE NO.', 'SENTENCE', 'DATE OF CONV.',
                'E.D.R', 'NAME OF A CHILD', 'AGE', 'SEX'
            ],
            'example_data': [
                ['01', 'TO 60/2024', 'ELIZA JOHN', 'V: CHIBWANA T/A: NCHILAMWERA D: THYOLO', 'F', '35',
                 'MURDER U/S 209 of P/C', 'HIGH COURT-BLANTYRE CC/NO.185/2020', '30 YEARS', '07-06-2024',
                 '06-06-2044', 'MIRACLE MUNYENGELA', '13 MONTHS', 'F'],
            ]
        },
        'pregnant_convicted': {
            'name': 'Pregnant Convicted Prisoners',
            'description': 'Return for convicted pregnant prisoners.',
            'columns': [
                'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE', 'OFFENSE',
                'COURT', 'SENTENCE', 'DOC', 'AMNESTY EARNED',
                'EXPECTED DATE OF RELEASE', 'GESTATION PERIOD', 'REMARKS'
            ],
            'example_data': [
                ['BT 9/25', 'JANE BANDA', 'V: CHIDE T/A: BVUMBWE D: THYOLO', 'M', '24',
                 'THEFT', 'F.G.M C/NO 785/25 MIDIMA', '30 MONTHS IHL', '02-07-2023',
                 'TWO-THREE MONTHS', '01-06-2024', '5 MONTHS', ''],
            ]
        },
        'pregnant_remand': {
            'name': 'Pregnant Remand Prisoners',
            'description': 'Return for pregnant remand (murder) prisoners.',
            'columns': [
                'S/N', 'REM.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT/CASE NO.', 'GESTATION PERIOD', 'REMARKS'
            ],
            'example_data': [
                ['01', '0102/24', 'YASINTA YOHANE', 'V: THABANI T/A: GOVATI D: MWANZA', 'F', '23',
                 'MURDER', 'HIGH COURT CC/NO.18/2025', '7 MONTHS', ''],
            ]
        },
        'children_remand': {
            'name': 'Children with Mothers on Remand',
            'description': 'Return for children accompanying their mothers on homicide and general remand.',
            'columns': [
                'S/N', 'PRI.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT CASE NO.', 'D.O.C', 'NAME OF A CHILD', 'AGE', 'SEX'
            ],
            'example_data': [
                ['1', 'TO 18/23', 'CHIKOND SAMERA', 'V: MAGOMBO T/A: NCHILAMWERA D: THYOLO', 'F', '22',
                 'MURDER', 'S.G.M THYOLO CC/NO.815/2023', '26-07-2023', 'STANCY MATIAS', '18 MONTHS', 'FEMALE'],
            ]
        },
        'foreigners_remand': {
            'name': 'Foreigners on Remand',
            'description': 'Return for foreign nationals on remand.',
            'columns': [
                'SER.NO', 'PRIS.NO', 'NAME', 'PARTICULARS', 'SEX', 'AGE',
                'OFFENSE', 'COURT AND CASE NO.', 'DATE OF ADMISSION',
                'ARRESTING AUTHORITY', 'PRE-TRIAL PERIOD', 'REMARKS'
            ],
            'example_data': [
                ['1', 'Mzu9/25', 'Tsedeke Girma', 'V: Hosana P: Hosana D: Hosana C: Zambia', 'M', '18',
                 'Illegal Entry', 'SGM C/No 14/25 Euthini', '22-02-2025', 'Kafukule Immigration', 'Waiting deportation', ''],
            ]
        },
    }

    @classmethod
    def create_default_templates(cls):
        """
        Create all default templates if they don't exist.
        """
        created_count = 0
        for category, definition in cls.TEMPLATE_DEFINITIONS.items():
            template, created = ReturnTemplate.objects.get_or_create(
                category=category,
                name=definition['name'],
                defaults={
                    'description': definition['description'],
                    'required_columns': definition['columns'],
                    'column_headers': definition['columns'],
                    'example_data': definition['example_data'],
                    'is_default': True,
                    'is_active': True,
                }
            )
            if created:
                created_count += 1
                logger.info(f"Created default template: {template.name}")

        return created_count

    @classmethod
    def get_template_by_category(cls, category):
        """Get template by category."""
        return ReturnTemplate.objects.filter(category=category, is_active=True).first()

    @classmethod
    def generate_template_csv(cls, category, period=None):
        """
        Generate a CSV template for a given category with headers and example data.
        """
        template = cls.get_template_by_category(category)
        if not template:
            raise ValidationError(f"No template found for category: {category}")

        columns = template.column_headers
        example_rows = template.example_data

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header row
        writer.writerow(columns)

        # Write example rows
        for example in example_rows:
            writer.writerow(example)

        return output.getvalue()

    @classmethod
    def generate_template_xlsx(cls, category, period=None):
        """
        Generate an Excel template for a given category with headers and example data.
        """
        template = cls.get_template_by_category(category)
        if not template:
            raise ValidationError(f"No template found for category: {category}")

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = template.name[:31]  # Excel sheet name max 31 chars

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        example_font = Font(italic=True, color="808080")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, header in enumerate(template.column_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write example rows
        for row_idx, example in enumerate(template.example_data, start=2):
            for col_idx, value in enumerate(example, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = example_font
                cell.border = border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        return wb


class ReturnProcessingService:
    """Service for processing return submissions."""

    # Column name mappings for standard return types
    COLUMN_MAPPINGS = {
        'convicted_inmates': {
            'ser_no': 'ser_no',
            'prisoner_number': ['pris_no', 'pri_no', 'prisoner_no', 'prn_no'],
            'name': ['name', 'full_name', 'prisoner_name'],
            'particulars': 'particulars',
            'sex': ['sex', 'gender'],
            'age': 'age',
            'offense': ['offense', 'offence', 'crime'],
            'court': ['court', 'court_name'],
            'case_no': ['case_no', 'case_number', 'court_case_no'],
            'sentence': ['sentence', 'sentence_months', 'term'],
            'date_of_committal': ['date_of_committal', 'committal_date', 'date_committed', 'doc'],
            'expiry_date_release_without_remission': [
                'expiry_date_without_remission', 'expiry_without_rem', 'release_date_no_rem'
            ],
            'expiry_date_release_with_remission': [
                'expiry_date_with_remission', 'expiry_with_rem', 'release_date_with_rem', 'edr'
            ],
        },
        'due_discharge': {
            'ser_no': 'ser_no',
            'prisoner_number': ['pris_no', 'pri_no', 'prisoner_no', 'prn_no'],
            'name': ['name', 'full_name', 'prisoner_name'],
            'particulars': 'particulars',
            'sex': ['sex', 'gender'],
            'age': 'age',
            'offense': ['offense', 'offence', 'crime'],
            'court': ['court', 'court_name'],
            'case_no': ['case_no', 'case_number', 'court_case_no'],
            'sentence': ['sentence', 'sentence_months', 'term'],
            'date_of_conviction': ['date_of_conviction', 'conviction_date', 'doc'],
            'amnesty_earned': ['amnesty_earned', 'amnesty', 'remission'],
            'due_date': ['due_date', 'discharge_date', 'release_date'],
            'remarks': ['remarks', 'notes', 'comment'],
        },
        'remand_murder': {
            'ser_no': 'ser_no',
            'prisoner_number': ['pris_no', 'pri_no', 'prisoner_no', 'prn_no'],
            'name': ['name', 'full_name', 'prisoner_name'],
            'particulars': 'particulars',
            'age': 'age',
            'sex': ['sex', 'gender'],
            'offense': ['offense', 'offence', 'crime'],
            'court_case_no': ['court_case_no', 'case_no', 'case_number'],
            'date_of_admission': ['date_of_admission', 'admission_date', 'doa'],
            'arresting_authority': ['arresting_authority', 'arrested_by'],
            'last_court_appearance': ['last_court_appearance', 'last_appearance'],
            'judge_name': ['judge_name', 'judge'],
            'status': ['status', 'remarks'],
        },
        'convicted_foreigners': {
            'ser_no': 'ser_no',
            'prisoner_number': ['pris_no', 'pri_no', 'prisoner_no', 'prn_no'],
            'name': ['name', 'full_name', 'prisoner_name'],
            'particulars': 'particulars',
            'sex': ['sex', 'gender'],
            'age': 'age',
            'offense': ['offense', 'offence', 'crime'],
            'court': ['court', 'court_name'],
            'sentence': ['sentence', 'sentence_months', 'term'],
            'date_of_conviction': ['date_of_conviction', 'conviction_date'],
            'amnesty_earned': ['amnesty_earned', 'amnesty', 'remission'],
            'expected_date_of_release': ['expected_date_of_release', 'edr', 'release_date'],
        },
        'general_remand': {
            'ser_no': 'ser_no',
            'name': ['name', 'full_name', 'prisoner_name'],
            'particulars': 'particulars',
            'sex': ['sex', 'gender'],
            'age': 'age',
            'offense': ['offense', 'offence', 'crime'],
            'court_case_no': ['court_case_no', 'case_no', 'case_number'],
            'date_of_admission': ['date_of_admission', 'admission_date', 'doa'],
            'arresting_authority': ['arresting_authority', 'arrested_by'],
            'last_court_appearance': ['last_court_appearance', 'last_appearance'],
            'case_status': ['case_status', 'status'],
        },
    }

    @classmethod
    def process_file(cls, submission: ReturnSubmission) -> Dict[str, Any]:
        """
        Process an uploaded return file.
        """
        result = {
            'success': False,
            'total_rows': 0,
            'valid_rows': 0,
            'error_rows': 0,
            'male_count': 0,
            'female_count': 0,
            'errors': []
        }

        try:
            # Read file based on extension
            file_content = submission.file.read()

            if submission.file.name.endswith('.csv'):
                rows = cls._parse_csv(file_content)
            elif submission.file.name.endswith(('.xlsx', '.xls')):
                rows = cls._parse_excel(file_content)
            else:
                raise ValidationError("Unsupported file format. Please upload CSV, XLSX, or XLS files.")

            # Check if first row is a header row
            # We need to be more careful here - only skip if it's clearly a header
            start_idx = 0
            if rows and cls._is_header_row(rows[0], submission.template):
                start_idx = 1

            # Filter out empty rows
            valid_rows_list = []
            for idx, row in enumerate(rows[start_idx:], start=start_idx + 1):
                # Skip completely empty rows
                if any(str(v).strip() for v in row.values() if v):
                    valid_rows_list.append((idx, row))

            result['total_rows'] = len(valid_rows_list)

            # Process rows
            for idx, row in valid_rows_list:
                try:
                    cleaned_data = cls._clean_row(row)
                    sex = cls._get_value(cleaned_data, 'sex').upper()

                    if sex == 'M':
                        result['male_count'] += 1
                    elif sex == 'F':
                        result['female_count'] += 1

                    # Get row number from the file (first column)
                    row_number = cls._parse_int(
                        cls._get_value(cleaned_data, 'no') or 
                        cls._get_value(cleaned_data, 'ser_no') or 
                        cls._get_value(cleaned_data, 's_no') or 
                        cls._get_value(cleaned_data, 's/n')
                    )
                    
                    ReturnData.objects.create(
                        submission=submission,
                        row_data=cleaned_data,
                        row_number=row_number or idx,  # Use row number from file, fallback to index
                        prisoner_number=cls._get_value(cleaned_data, 'prisoner_number'),
                        name=cls._get_value(cleaned_data, 'name'),
                        sex=sex,
                        age=cls._parse_int(cls._get_value(cleaned_data, 'age')),
                        offense=cls._get_value(cleaned_data, 'offense'),
                        court_case_number=cls._get_value(cleaned_data, 'court_case_no') or cls._get_value(cleaned_data, 'case_no'),
                        sentence=cls._get_value(cleaned_data, 'sentence'),
                        date_of_committal=cls._parse_date(cls._get_value(cleaned_data, 'date_of_committal') or cls._get_value(cleaned_data, 'date_of_admission')),
                        expiry_date=cls._parse_date(cls._get_value(cleaned_data, 'expiry_date_release_with_remission') or cls._get_value(cleaned_data, 'expected_date_of_release')),
                    )
                    result['valid_rows'] += 1
                except Exception as e:
                    result['error_rows'] += 1
                    result['errors'].append(f"Row {idx}: {str(e)}")

            # Update submission
            submission.row_count = result['total_rows']
            submission.error_count = result['error_rows']
            submission.total_male = result['male_count']
            submission.total_female = result['female_count']
            submission.status = 'imported'
            submission.processed_at = timezone.now()
            submission.processed_by = submission.submitted_by
            submission.error_log = '\n'.join(result['errors'][:100]) if result['errors'] else ''
            submission.save()

            result['success'] = True
            return result

        except Exception as e:
            logger.error(f"Error processing return file: {str(e)}")
            submission.status = 'rejected'
            submission.error_log = str(e)
            submission.save()
            return {
                'success': False,
                'errors': [str(e)]
            }

    @classmethod
    def _is_header_row(cls, row: Dict[str, str], template: ReturnTemplate) -> bool:
        """Check if row appears to be a header row."""
        if not row:
            return False

        # Get values from the row
        values = [str(v).strip().upper() for v in row.values() if v]
        
        if not values:
            return False

        # Check if the first few values look like headers
        headers = template.column_headers
        header_upper = [h.upper() for h in headers]
        
        # Count how many values match header names
        matches = sum(1 for value in values[:5] if value in header_upper)
        
        # Also check if values look like headers (contain common header keywords)
        header_keywords = ['NO', 'NAME', 'SEX', 'AGE', 'OFFENSE', 'COURT', 'SENTENCE', 'DATE', 'PRIS', 'PARTICULARS']
        keyword_matches = sum(1 for value in values[:5] if any(keyword in value for keyword in header_keywords))
        
        # If values don't look like data (not numbers, not names with spaces), treat as header
        # A data row typically has: a number, a name, a sex (M/F), an age (number)
        first_value = values[0] if values else ''
        
        # If first value is "NO" or contains "PRIS" or "SER", it's a header
        if first_value in ['NO', 'S/NO', 'SER.NO', 'S/N', 'SER.NO'] or 'PRIS' in first_value:
            return True
            
        # Check if values match header names
        if matches >= 2:
            return True
            
        return False

    @classmethod
    def _parse_csv(cls, file_content: bytes) -> List[Dict[str, str]]:
        """Parse CSV file content."""
        decoded = file_content.decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(row) for row in reader if any(row.values())]

    @classmethod
    def _parse_excel(cls, file_content: bytes) -> List[Dict[str, str]]:
        """Parse Excel file content."""
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            sheet = wb.active

            rows = []
            headers = None
            for row in sheet.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    continue

                if any(cell is not None and str(cell).strip() for cell in row):
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            row_dict[header] = str(row[idx]) if row[idx] is not None else ''
                    rows.append(row_dict)

            wb.close()
            return rows

        except ImportError:
            logger.error("openpyxl not installed. Cannot parse Excel files.")
            raise ValidationError("Excel parsing requires openpyxl. Please install it or use CSV format.")
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise ValidationError(f"Could not parse Excel file: {str(e)}")

    @classmethod
    def _clean_row(cls, row: Dict[str, str]) -> Dict[str, str]:
        """Clean a row by standardizing keys and removing empty values."""
        cleaned = {}
        for key, value in row.items():
            if key:
                cleaned[str(key).strip().lower().replace(' ', '_')] = str(value).strip() if value else ''
        return cleaned

    @classmethod
    def _get_value(cls, data: Dict[str, Any], field: str) -> str:
        """Get value from cleaned data dictionary."""
        if not data:
            return ''
        return str(data.get(field, '') or '').strip()

    @classmethod
    def _parse_int(cls, value: str) -> Optional[int]:
        """Parse integer from string."""
        if not value or value == 'None':
            return None
        try:
            return int(float(value.replace(',', '')))
        except (ValueError, TypeError):
            return None

    @classmethod
    def _parse_date(cls, value: str) -> Optional[Any]:
        """Parse date from string."""
        if not value or value == 'None':
            return None

        # Try multiple date formats
        date_formats = [
            '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
            '%Y/%m/%d', '%d.%m.%Y', '%b %d, %Y', '%d %b %Y'
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        return None


class ReturnReportService:
    """Service for generating return reports and summaries."""

    @classmethod
    def generate_regional_summary(cls, period: str, user) -> Dict[str, Any]:
        """
        Generate a regional summary for a given period.
        Aggregates data from all submissions for that period.
        """
        summary = {
            'period': period,
            'total_submissions': 0,
            'total_records': 0,
            'total_male': 0,
            'total_female': 0,
            'by_category': {},
            'by_station': {},
            'by_region': {},
        }

        # Get all submissions for the period
        submissions = ReturnSubmission.objects.filter(period=period)

        summary['total_submissions'] = submissions.count()
        summary['total_records'] = sum(s.row_count for s in submissions)
        summary['total_male'] = sum(s.total_male for s in submissions)
        summary['total_female'] = sum(s.total_female for s in submissions)

        # Aggregate by category
        for submission in submissions:
            category = submission.template.category
            if category not in summary['by_category']:
                summary['by_category'][category] = {
                    'count': 0,
                    'records': 0,
                    'stations': set(),
                    'male': 0,
                    'female': 0
                }
            summary['by_category'][category]['count'] += 1
            summary['by_category'][category]['records'] += submission.row_count
            summary['by_category'][category]['male'] += submission.total_male
            summary['by_category'][category]['female'] += submission.total_female
            summary['by_category'][category]['stations'].add(submission.prison_station.name)

        # Aggregate by station
        for submission in submissions:
            station_name = submission.prison_station.name
            if station_name not in summary['by_station']:
                summary['by_station'][station_name] = {
                    'count': 0,
                    'records': 0,
                    'male': 0,
                    'female': 0,
                    'categories': []
                }
            summary['by_station'][station_name]['count'] += 1
            summary['by_station'][station_name]['records'] += submission.row_count
            summary['by_station'][station_name]['male'] += submission.total_male
            summary['by_station'][station_name]['female'] += submission.total_female
            summary['by_station'][station_name]['categories'].append(submission.template.get_category_display())

        # Aggregate by region
        for submission in submissions:
            region = submission.prison_station.get_region_display()
            if region not in summary['by_region']:
                summary['by_region'][region] = {
                    'stations': 0,
                    'submissions': 0,
                    'records': 0,
                    'male': 0,
                    'female': 0
                }
            summary['by_region'][region]['submissions'] += 1
            summary['by_region'][region]['records'] += submission.row_count
            summary['by_region'][region]['male'] += submission.total_male
            summary['by_region'][region]['female'] += submission.total_female

        # Count unique stations per region
        regions = PrisonStation.objects.values('region').annotate(
            station_count=Count('id')
        )
        for region_data in regions:
            region_display = dict(PrisonStation.REGION_CHOICES).get(region_data['region'], region_data['region'])
            if region_display in summary['by_region']:
                summary['by_region'][region_display]['stations'] = region_data['station_count']

        # Convert sets to lists for JSON serialization
        for category_data in summary['by_category'].values():
            category_data['stations'] = sorted(category_data['stations'])

        # Create/save RegionalReturnSummary
        for region, data in summary['by_region'].items():
            RegionalReturnSummary.objects.update_or_create(
                category='all',
                region=region,
                period=period,
                period_type='monthly',
                defaults={
                    'total_records': data['records'],
                    'male_count': data['male'],
                    'female_count': data['female'],
                    'additional_data': data,
                    'generated_by': user
                }
            )

        return summary

    @classmethod
    def generate_monthly_status_report(cls, year: int, month: int) -> Dict[str, Any]:
        """
        Generate a monthly status report for all stations.
        Shows which stations have submitted which returns.
        """
        period = f"{year}-{month:02d}"
        
        report = {
            'year': year,
            'month': month,
            'period': period,
            'total_stations': 0,
            'stations_submitted_all': 0,
            'stations_missing': [],
            'by_station': {},
        }

        # Get all active stations
        stations = PrisonStation.objects.all()
        report['total_stations'] = stations.count()

        # Get all active templates
        templates = ReturnTemplate.objects.filter(is_active=True)
        total_templates = templates.count()

        # Get all statuses for this period
        statuses = StationReturnStatus.objects.filter(
            year=year,
            month=month
        ).select_related('prison_station', 'template')

        # Build status map
        status_map = {}
        for status in statuses:
            key = (status.prison_station_id, status.template_id)
            status_map[key] = status

        # Build report by station
        for station in stations:
            station_data = {
                'station': station,
                'submitted_count': 0,
                'approved_count': 0,
                'pending_count': 0,
                'missing_count': 0,
                'templates': [],
                'has_all': False,
            }

            for template in templates:
                status = status_map.get((station.id, template.id))
                template_status = {
                    'template': template,
                    'status': status.status if status else 'not_submitted',
                    'submitted_at': status.submitted_at if status else None,
                    'approved_at': status.approved_at if status else None,
                }

                if status:
                    if status.status in ['submitted', 'approved']:
                        station_data['submitted_count'] += 1
                    if status.status == 'approved':
                        station_data['approved_count'] += 1
                    if status.status in ['pending', 'not_submitted']:
                        station_data['pending_count'] += 1
                else:
                    station_data['missing_count'] += 1

                station_data['templates'].append(template_status)

            station_data['has_all'] = station_data['submitted_count'] == total_templates
            report['by_station'][station.name] = station_data

            if station_data['has_all']:
                report['stations_submitted_all'] += 1
            else:
                report['stations_missing'].append(station.name)

        return report


class MonthlySubmissionService:
    """
    Service for managing monthly submission tracking.
    """

    @classmethod
    def initialize_monthly_tracking(cls, year: int, month: int) -> Dict[str, Any]:
        """
        Initialize monthly tracking for all stations and templates.
        Creates StationReturnStatus records for all stations/templates for the period.
        """
        period = f"{year}-{month:02d}"
        stations = PrisonStation.objects.all()
        templates = ReturnTemplate.objects.filter(is_active=True)

        created_count = 0
        for station in stations:
            for template in templates:
                obj, created = StationReturnStatus.objects.get_or_create(
                    prison_station=station,
                    template=template,
                    period=period,
                    defaults={
                        'year': year,
                        'month': month,
                        'status': 'not_submitted',
                    }
                )
                if created:
                    created_count += 1

        return {
            'created_count': created_count,
            'total_expected': stations.count() * templates.count(),
            'period': period,
        }

    @classmethod
    def update_station_status(cls, submission: ReturnSubmission) -> None:
        """
        Update station status when a submission is made.
        """
        period = submission.period

        status, created = StationReturnStatus.objects.update_or_create(
            prison_station=submission.prison_station,
            template=submission.template,
            period=period,
            defaults={
                'year': submission.year,
                'month': submission.month,
                'status': 'submitted',
                'submission': submission,
                'submitted_at': timezone.now(),
            }
        )

        # Update MonthlySubmissionTracker
        cls._update_monthly_tracker(submission.prison_station, submission.year, submission.month)

    @classmethod
    def approve_station_status(cls, submission: ReturnSubmission) -> None:
        """
        Update station status when a submission is approved.
        """
        period = submission.period

        status, created = StationReturnStatus.objects.update_or_create(
            prison_station=submission.prison_station,
            template=submission.template,
            period=period,
            defaults={
                'year': submission.year,
                'month': submission.month,
                'status': 'approved',
                'submission': submission,
                'submitted_at': submission.submitted_at,
                'approved_at': timezone.now(),
            }
        )

        # Update MonthlySubmissionTracker
        cls._update_monthly_tracker(submission.prison_station, submission.year, submission.month)

    @classmethod
    def reject_station_status(cls, submission: ReturnSubmission, reason: str = '') -> None:
        """
        Update station status when a submission is rejected.
        """
        period = submission.period

        status, created = StationReturnStatus.objects.update_or_create(
            prison_station=submission.prison_station,
            template=submission.template,
            period=period,
            defaults={
                'year': submission.year,
                'month': submission.month,
                'status': 'rejected',
                'submission': submission,
                'submitted_at': submission.submitted_at,
                'rejected_at': timezone.now(),
            }
        )

        # Update MonthlySubmissionTracker
        cls._update_monthly_tracker(submission.prison_station, submission.year, submission.month)

    @classmethod
    def _update_monthly_tracker(cls, station: PrisonStation, year: int, month: int) -> None:
        """
        Update or create MonthlySubmissionTracker for a station.
        """
        period = f"{year}-{month:02d}"

        # Get all statuses for this station/period
        statuses = StationReturnStatus.objects.filter(
            prison_station=station,
            year=year,
            month=month
        )

        total_required = statuses.count()
        total_submitted = statuses.filter(status__in=['submitted', 'approved']).count()
        total_approved = statuses.filter(status='approved').count()
        total_rejected = statuses.filter(status='rejected').count()
        total_pending = statuses.filter(status='pending').count()

        # Build template statuses dict
        template_statuses = {}
        for status in statuses:
            template_statuses[status.template.category] = status.status

        tracker, created = MonthlySubmissionTracker.objects.update_or_create(
            prison_station=station,
            year=year,
            month=month,
            period=period,
            defaults={
                'total_required': total_required,
                'total_submitted': total_submitted,
                'total_approved': total_approved,
                'total_rejected': total_rejected,
                'total_pending': total_pending,
                'template_statuses': template_statuses,
                'is_complete': total_submitted == total_required,
            }
        )

        return tracker